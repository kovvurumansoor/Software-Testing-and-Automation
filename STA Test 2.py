from faker import Faker
from openpyxl import Workbook

# Create workbook and worksheet
wb = Workbook()
ws = wb.active

# Create Faker object
fake_data = Faker()

# Add headings
ws.cell(row=1, column=1).value = "Name"
ws.cell(row=1, column=2).value = "Email"

# Generate fake data
for i in range(2, 12):
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()

# Save workbook
wb.save("fake_data.xlsx")

print("Excel file created successfully!")

if ws.max_row > 1:
    print("✅ Data uploaded successfully!")
else:
    print("❌ Data upload failed!")
