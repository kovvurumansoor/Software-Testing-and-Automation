from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# Read Excel
wb = load_workbook("ContactData.xlsx")
ws = wb.active

name = ws["A2"].value
email = ws["B2"].value
subject = ws["C2"].value
message = ws["D2"].value

# Open Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Change this path to your HTML file
driver.get(r"C:\Users\Lenovo\PycharmProjects\Software Testing and Automation\Contact Test 3.html")

driver.maximize_window()

# Fill the form
driver.find_element(By.ID, "name").send_keys(name)
driver.find_element(By.ID, "email").send_keys(email)
driver.find_element(By.ID, "subject").send_keys(subject)
driver.find_element(By.ID, "message").send_keys(message)

# Click Submit automatically
driver.find_element(By.ID, "submit").click()

# Verify success message
success = driver.find_element(By.ID, "success-message").text

if success == "Contact form submitted successfully!":
    print("✅ Test Passed")
else:
    print("❌ Test Failed")

input("Press Enter to close the browser...")
driver.quit()